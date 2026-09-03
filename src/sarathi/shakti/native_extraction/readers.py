"""Concrete format readers for Shruti Native Extraction."""

from __future__ import annotations

import csv
import io
import xml.etree.ElementTree as ET
from typing import Any
from zipfile import BadZipFile

import charset_normalizer
import openpyxl
import polars as pl
import pymupdf
import python_calamine
import xlrd
from bs4 import BeautifulSoup

from sarathi.dosh import DoshError, FailureCode
from sarathi.sankalpa import (
    CanonicalDocument,
    PageData,
    ProvenanceRecord,
    TableData,
    TextSpan,
    WarningRecord,
)

_STAGE_NAME = "read_native"
_PLUGIN_ID = "shakti.native_extraction"
_CAPABILITY_ID = "read_native"


def read_pdf(
    data: bytes,
    input_id: str,
) -> tuple[CanonicalDocument, tuple[ProvenanceRecord, ...], tuple[WarningRecord, ...]]:
    """Extract native text, spans, and embedded tables from PDF via PyMuPDF."""
    doc = pymupdf.open(stream=data, filetype="pdf")
    pages: list[PageData] = []
    provenances: list[ProvenanceRecord] = []
    warnings: list[WarningRecord] = []
    full_text_parts: list[str] = []
    all_doc_tables: list[TableData] = []

    try:
        total_pages = len(doc)
        for page_idx in range(total_pages):
            page_num = page_idx + 1
            page = doc[page_idx]
            page_text = page.get_text("text").strip()
            if page_text:
                full_text_parts.append(page_text)

            # Extract block spans
            spans: list[TextSpan] = []
            blocks = page.get_text("blocks")
            for b in blocks:
                if len(b) >= 5:
                    x0, y0, x1, y1, text = b[0], b[1], b[2], b[3], b[4]
                    if isinstance(text, str) and text.strip():
                        spans.append(
                            TextSpan(
                                text=text.strip(),
                                bounding_box=(float(x0), float(y0), float(x1), float(y1)),
                            )
                        )

            # Extract native vector tables if present
            page_tables: list[TableData] = []
            tabs = None
            try:
                tabs = page.find_tables()
            except (pymupdf.FileDataError, ValueError):
                warnings.append(
                    WarningRecord(
                        code="PDF_TABLE_DETECTION_SKIPPED",
                        message="Vector table extraction skipped for page.",
                        stage=_STAGE_NAME,
                    )
                )

            if tabs and len(tabs.tables) > 0:
                for t_idx, tab in enumerate(tabs.tables, 1):
                    extracted_rows = tab.extract()
                    if extracted_rows and len(extracted_rows) > 0:
                        headers = tuple(str(h or "") for h in extracted_rows[0])
                        data_rows = tuple(tuple(val for val in row) for row in extracted_rows[1:])
                        t_obj = TableData(
                            name=f"Page_{page_num}_Table_{t_idx}",
                            headers=headers,
                            rows=data_rows,
                        )
                        page_tables.append(t_obj)
                        all_doc_tables.append(t_obj)

            pages.append(
                PageData(
                    page_number=page_num,
                    text=page_text,
                    spans=tuple(spans),
                    tables=tuple(page_tables),
                )
            )

            provenances.append(
                ProvenanceRecord(
                    source_input_id=input_id,
                    stage=_STAGE_NAME,
                    plugin_id=_PLUGIN_ID,
                    capability_id=_CAPABILITY_ID,
                    page_number=page_num,
                    evidence={
                        "reader": "pymupdf",
                        "page_count": total_pages,
                        "has_native_text": bool(page_text),
                        "table_count": len(page_tables),
                    },
                )
            )
    finally:
        doc.close()

    canonical_doc = CanonicalDocument(
        document_id=f"doc-{input_id}",
        source_input_id=input_id,
        pages=tuple(pages),
        tables=tuple(all_doc_tables),
        text="\n\n".join(full_text_parts),
        detected_type="pdf",
    )
    return canonical_doc, tuple(provenances), tuple(warnings)


def read_xlsx(
    data: bytes,
    input_id: str,
) -> tuple[CanonicalDocument, tuple[ProvenanceRecord, ...], tuple[WarningRecord, ...]]:
    """Extract all sheets from XLSX/XLSM using python-calamine with openpyxl fallback."""
    tables: list[TableData] = []
    provenances: list[ProvenanceRecord] = []
    warnings: list[WarningRecord] = []
    reader_used = "python-calamine"

    try:
        wb = python_calamine.CalamineWorkbook.from_object(io.BytesIO(data))
        for sheet_name in wb.sheet_names:
            sheet = wb.get_sheet_by_name(sheet_name)
            raw_rows = sheet.to_python()
            if raw_rows and len(raw_rows) > 0:
                headers = tuple(str(col or "") for col in raw_rows[0])
                data_rows = tuple(tuple(cell for cell in row) for row in raw_rows[1:])
            else:
                headers = ()
                data_rows = ()

            tables.append(
                TableData(
                    name=sheet_name,
                    headers=headers,
                    rows=data_rows,
                )
            )
            provenances.append(
                ProvenanceRecord(
                    source_input_id=input_id,
                    stage=_STAGE_NAME,
                    plugin_id=_PLUGIN_ID,
                    capability_id=_CAPABILITY_ID,
                    evidence={"reader": "python-calamine", "sheet": sheet_name, "row_count": len(raw_rows)},
                )
            )
    except (python_calamine.CalamineError, BadZipFile):
        # Fallback to openpyxl
        reader_used = "openpyxl"
        warnings.append(
            WarningRecord(
                code="CALAMINE_FALLBACK",
                message="python-calamine failed on workbook, fell back to openpyxl reader.",
                stage=_STAGE_NAME,
            )
        )
        wb_px = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        try:
            for ws in wb_px.worksheets:
                rows_list = list(ws.iter_rows(values_only=True))
                if rows_list and len(rows_list) > 0:
                    headers = tuple(str(c or "") for c in rows_list[0])
                    data_rows = tuple(tuple(c for c in row) for row in rows_list[1:])
                else:
                    headers = ()
                    data_rows = ()

                tables.append(
                    TableData(
                        name=ws.title,
                        headers=headers,
                        rows=data_rows,
                    )
                )
                provenances.append(
                    ProvenanceRecord(
                        source_input_id=input_id,
                        stage=_STAGE_NAME,
                        plugin_id=_PLUGIN_ID,
                        capability_id=_CAPABILITY_ID,
                        evidence={"reader": "openpyxl", "sheet": ws.title, "row_count": len(rows_list)},
                    )
                )
        finally:
            wb_px.close()

    # Check for active filters and hidden rows across worksheets
    filter_detected = False
    filter_details: list[str] = []
    try:
        wb_filter_check = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=False)
        try:
            for ws_c in wb_filter_check.worksheets:
                has_af = bool(ws_c.auto_filter and ws_c.auto_filter.ref)
                has_hidden = any(ws_c.row_dimensions[r].hidden for r in ws_c.row_dimensions)
                if has_af or has_hidden:
                    filter_detected = True
                    af_ref = ws_c.auto_filter.ref if has_af else "none"
                    filter_details.append(f"{ws_c.title}(auto_filter={af_ref}, hidden_rows={has_hidden})")
                    warnings.append(
                        WarningRecord(
                            code="EXCEL_FILTER_APPLIED",
                            message=f"Worksheet '{ws_c.title}' has an AutoFilter ({af_ref}) or hidden rows. All rows are extracted.",
                            stage=_STAGE_NAME,
                        )
                    )
        finally:
            wb_filter_check.close()
    except Exception:
        pass

    # Build composite text from table headers and cells for downstream processing
    composite_lines: list[str] = []
    for t in tables:
        if t.name:
            composite_lines.append(f"--- Sheet: {t.name} ---")
        if t.headers:
            composite_lines.append(" | ".join(str(h) for h in t.headers if str(h).strip()))
        for row in t.rows:
            row_str = " | ".join(str(c) for c in row if c is not None and str(c).strip())
            if row_str:
                composite_lines.append(row_str)
    full_text = "\n".join(composite_lines)

    meta: dict[str, Any] = {"reader": reader_used}
    if filter_detected:
        meta["filter_applied"] = True
        meta["filter_details"] = "; ".join(filter_details)

    canonical_doc = CanonicalDocument(
        document_id=f"doc-{input_id}",
        source_input_id=input_id,
        text=full_text,
        tables=tuple(tables),
        detected_type="xlsx",
        metadata=meta,
    )
    return canonical_doc, tuple(provenances), tuple(warnings)


def read_xls_legacy(
    data: bytes,
    input_id: str,
) -> tuple[CanonicalDocument, tuple[ProvenanceRecord, ...], tuple[WarningRecord, ...]]:
    """Extract all sheets from legacy BIFF .xls using xlrd."""
    tables: list[TableData] = []
    provenances: list[ProvenanceRecord] = []
    warnings: list[WarningRecord] = []

    rb = xlrd.open_workbook(file_contents=data)
    for sheet_idx in range(rb.nsheets):
        sheet = rb.sheet_by_index(sheet_idx)
        rows_list: list[tuple[Any, ...]] = [tuple(sheet.row_values(r)) for r in range(sheet.nrows)]
        if rows_list and len(rows_list) > 0:
            headers = tuple(str(c or "") for c in rows_list[0])
            data_rows = tuple(rows_list[1:])
        else:
            headers = ()
            data_rows = ()

        tables.append(
            TableData(
                name=sheet.name,
                headers=headers,
                rows=data_rows,
            )
        )
        provenances.append(
            ProvenanceRecord(
                source_input_id=input_id,
                stage=_STAGE_NAME,
                plugin_id=_PLUGIN_ID,
                capability_id=_CAPABILITY_ID,
                evidence={"reader": "xlrd", "sheet": sheet.name, "row_count": len(rows_list)},
            )
        )

    canonical_doc = CanonicalDocument(
        document_id=f"doc-{input_id}",
        source_input_id=input_id,
        tables=tuple(tables),
        detected_type="xls_legacy",
    )
    return canonical_doc, tuple(provenances), tuple(warnings)


def read_html_table(
    data: bytes,
    input_id: str,
) -> tuple[CanonicalDocument, tuple[ProvenanceRecord, ...], tuple[WarningRecord, ...]]:
    """Extract tables or text from HTML markup (including disguised .xls) via BeautifulSoup."""
    tables: list[TableData] = []
    provenances: list[ProvenanceRecord] = []
    warnings: list[WarningRecord] = []

    match = charset_normalizer.from_bytes(data).best()
    encoding = match.encoding if match and match.encoding else "utf-8"
    try:
        html_text = data.decode(encoding)
    except (UnicodeDecodeError, LookupError):
        html_text = data.decode("utf-8", errors="replace")

    soup = BeautifulSoup(html_text, "html.parser")
    html_tables = soup.find_all("table")

    if html_tables and len(html_tables) > 0:
        for t_idx, t_tag in enumerate(html_tables, 1):
            rows: list[tuple[Any, ...]] = []
            headers: tuple[str, ...] = ()
            for tr_idx, tr in enumerate(t_tag.find_all("tr")):
                th_cells = [th.get_text(strip=True) for th in tr.find_all("th")]
                td_cells = [td.get_text(strip=True) for td in tr.find_all(["td", "th"])]
                if tr_idx == 0 and th_cells:
                    headers = tuple(th_cells)
                elif tr_idx == 0 and td_cells and not headers:
                    headers = tuple(td_cells)
                else:
                    if td_cells:
                        rows.append(tuple(td_cells))

            table_name = t_tag.get("id") or t_tag.get("name") or f"Table_{t_idx}"
            tables.append(
                TableData(
                    name=str(table_name),
                    headers=headers,
                    rows=tuple(rows),
                )
            )
            provenances.append(
                ProvenanceRecord(
                    source_input_id=input_id,
                    stage=_STAGE_NAME,
                    plugin_id=_PLUGIN_ID,
                    capability_id=_CAPABILITY_ID,
                    evidence={"reader": "beautifulsoup4", "table_name": str(table_name), "row_count": len(rows)},
                )
            )
        pages = ()
        doc_text = ""
    else:
        # Fallback to plain text extracted from body
        doc_text = soup.get_text(separator="\n", strip=True)
        pages = (PageData(page_number=1, text=doc_text),)
        provenances.append(
            ProvenanceRecord(
                source_input_id=input_id,
                stage=_STAGE_NAME,
                plugin_id=_PLUGIN_ID,
                capability_id=_CAPABILITY_ID,
                page_number=1,
                evidence={"reader": "beautifulsoup4", "type": "html_text"},
            )
        )

    canonical_doc = CanonicalDocument(
        document_id=f"doc-{input_id}",
        source_input_id=input_id,
        pages=pages,
        tables=tuple(tables),
        text=doc_text,
        detected_type="html_table",
    )
    return canonical_doc, tuple(provenances), tuple(warnings)


def read_spreadsheet_ml(
    data: bytes,
    input_id: str,
) -> tuple[CanonicalDocument, tuple[ProvenanceRecord, ...], tuple[WarningRecord, ...]]:
    """Extract XML spreadsheets (2003 XML format) using ElementTree."""
    tables: list[TableData] = []
    provenances: list[ProvenanceRecord] = []
    warnings: list[WarningRecord] = []

    root = ET.fromstring(data)
    # Strip namespace for robust tag matching
    ns = ""
    if root.tag.startswith("{"):
        ns = root.tag.split("}")[0] + "}"

    worksheets = root.findall(f"{ns}Worksheet")
    for ws_idx, ws in enumerate(worksheets, 1):
        ws_name = ws.attrib.get(f"{ns}Name", f"Sheet_{ws_idx}")
        table_elem = ws.find(f"{ns}Table")
        rows: list[tuple[Any, ...]] = []
        headers: tuple[str, ...] = ()

        if table_elem is not None:
            for r_idx, row_elem in enumerate(table_elem.findall(f"{ns}Row")):
                cell_vals: list[str] = []
                for cell_elem in row_elem.findall(f"{ns}Cell"):
                    data_elem = cell_elem.find(f"{ns}Data")
                    val = data_elem.text if data_elem is not None and data_elem.text else ""
                    cell_vals.append(val.strip())

                if r_idx == 0:
                    headers = tuple(cell_vals)
                else:
                    rows.append(tuple(cell_vals))

        tables.append(
            TableData(
                name=ws_name,
                headers=headers,
                rows=tuple(rows),
            )
        )
        provenances.append(
            ProvenanceRecord(
                source_input_id=input_id,
                stage=_STAGE_NAME,
                plugin_id=_PLUGIN_ID,
                capability_id=_CAPABILITY_ID,
                evidence={"reader": "elementtree", "sheet": ws_name, "row_count": len(rows)},
            )
        )

    canonical_doc = CanonicalDocument(
        document_id=f"doc-{input_id}",
        source_input_id=input_id,
        tables=tuple(tables),
        detected_type="spreadsheet_ml",
    )
    return canonical_doc, tuple(provenances), tuple(warnings)


def read_csv_or_text(
    data: bytes,
    input_id: str,
) -> tuple[CanonicalDocument, tuple[ProvenanceRecord, ...], tuple[WarningRecord, ...]]:
    """Extract CSV or plain text using polars / stdlib csv with charset-normalizer encoding detection."""
    match = charset_normalizer.from_bytes(data).best()
    encoding = match.encoding if match and match.encoding else "utf-8"
    try:
        text_content = data.decode(encoding)
    except (UnicodeDecodeError, LookupError):
        text_content = data.decode("utf-8", errors="replace")

    provenances: list[ProvenanceRecord] = []
    warnings: list[WarningRecord] = []
    tables: list[TableData] = []
    pages: list[PageData] = []

    # Attempt tabular parsing via polars
    parsed_tabular = False
    try:
        df = pl.read_csv(io.BytesIO(data), encoding=encoding, infer_schema_length=100)
        if len(df.columns) > 1 or len(df) > 0:
            headers = tuple(df.columns)
            rows = tuple(tuple(val for val in row) for row in df.iter_rows())
            tables.append(TableData(name="default", headers=headers, rows=rows))
            provenances.append(
                ProvenanceRecord(
                    source_input_id=input_id,
                    stage=_STAGE_NAME,
                    plugin_id=_PLUGIN_ID,
                    capability_id=_CAPABILITY_ID,
                    evidence={"reader": "polars", "encoding": encoding, "row_count": len(rows)},
                )
            )
            parsed_tabular = True
    except (pl.exceptions.PolarsError, csv.Error, UnicodeDecodeError):
        # Fallback to stdlib csv
        try:
            sample = text_content[:2048]
            dialect = csv.Sniffer().sniff(sample)
            reader = csv.reader(io.StringIO(text_content), dialect)
            all_rows = list(reader)
            if all_rows and len(all_rows[0]) > 1:
                headers = tuple(all_rows[0])
                rows = tuple(tuple(row) for row in all_rows[1:])
                tables.append(TableData(name="default", headers=headers, rows=rows))
                provenances.append(
                    ProvenanceRecord(
                        source_input_id=input_id,
                        stage=_STAGE_NAME,
                        plugin_id=_PLUGIN_ID,
                        capability_id=_CAPABILITY_ID,
                        evidence={"reader": "csv", "encoding": encoding, "row_count": len(rows)},
                    )
                )
                parsed_tabular = True
        except (csv.Error, UnicodeDecodeError):
            pass

    if not parsed_tabular:
        # Plain text
        pages.append(PageData(page_number=1, text=text_content.strip()))
        provenances.append(
            ProvenanceRecord(
                source_input_id=input_id,
                stage=_STAGE_NAME,
                plugin_id=_PLUGIN_ID,
                capability_id=_CAPABILITY_ID,
                page_number=1,
                evidence={"reader": "charset_normalizer", "encoding": encoding, "char_count": len(text_content)},
            )
        )

    canonical_doc = CanonicalDocument(
        document_id=f"doc-{input_id}",
        source_input_id=input_id,
        pages=tuple(pages),
        tables=tuple(tables),
        text=text_content.strip() if not parsed_tabular else "",
        detected_type="csv_or_text",
    )
    return canonical_doc, tuple(provenances), tuple(warnings)


def read_docx(
    data: bytes,
    input_id: str,
) -> tuple[CanonicalDocument, tuple[ProvenanceRecord, ...], tuple[WarningRecord, ...]]:
    """Extract full text, paragraphs, and tables from a DOCX (OpenXML) document."""
    import xml.etree.ElementTree as ET
    import zipfile

    tables: list[TableData] = []
    paragraphs: list[str] = []
    provenances: list[ProvenanceRecord] = []
    warnings: list[WarningRecord] = []

    _W_NAMESPACE = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
    _P_TAG = f"{_W_NAMESPACE}p"
    _R_TAG = f"{_W_NAMESPACE}r"
    _T_TAG = f"{_W_NAMESPACE}t"
    _TBL_TAG = f"{_W_NAMESPACE}tbl"
    _TR_TAG = f"{_W_NAMESPACE}tr"
    _TC_TAG = f"{_W_NAMESPACE}tc"

    with zipfile.ZipFile(io.BytesIO(data)) as zf:
        if "word/document.xml" not in zf.namelist():
            raise DoshError(
                code=FailureCode.UNSUPPORTED,
                message="DOCX file is missing required word/document.xml.",
            )
        doc_xml = zf.read("word/document.xml")
        tree = ET.fromstring(doc_xml)

        body = tree.find(f"{_W_NAMESPACE}body")
        if body is not None:
            tbl_count = 0
            for elem in body:
                if elem.tag == _P_TAG:
                    p_text = "".join(t.text for t in elem.iter(_T_TAG) if t.text)
                    if p_text.strip():
                        paragraphs.append(p_text.strip())
                elif elem.tag == _TBL_TAG:
                    tbl_count += 1
                    raw_table_rows: list[tuple[str, ...]] = []
                    for tr in elem.findall(_TR_TAG):
                        row_cells: list[str] = []
                        for tc in tr.findall(_TC_TAG):
                            tc_text = "".join(t.text for t in tc.iter(_T_TAG) if t.text).strip()
                            row_cells.append(tc_text)
                        if any(row_cells):
                            raw_table_rows.append(tuple(row_cells))

                    if raw_table_rows:
                        headers = raw_table_rows[0]
                        data_rows = tuple(raw_table_rows[1:])
                        tables.append(
                            TableData(
                                name=f"Table_{tbl_count}",
                                headers=headers,
                                rows=data_rows,
                            )
                        )
                        for r in raw_table_rows:
                            paragraphs.append(" | ".join(r))

    provenances.append(
        ProvenanceRecord(
            source_input_id=input_id,
            stage=_STAGE_NAME,
            plugin_id=_PLUGIN_ID,
            capability_id=_CAPABILITY_ID,
            evidence={
                "reader": "docx_openxml",
                "paragraph_count": len(paragraphs),
                "table_count": len(tables),
            },
        )
    )

    full_text = "\n".join(paragraphs)
    canonical_doc = CanonicalDocument(
        document_id=f"doc-{input_id}",
        source_input_id=input_id,
        text=full_text,
        tables=tuple(tables),
        detected_type="docx",
        metadata={"reader": "docx_openxml"},
    )
    return canonical_doc, tuple(provenances), tuple(warnings)
