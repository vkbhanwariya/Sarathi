"""Consolidator and Canonical Output Exporter for Bank Statements.

Produces:
1. Consolidated_Bank_Statement.parquet (Polars persistent machine analysis source preserving exact Decimal precision)
2. Consolidated_Bank_Statement.xlsx (openpyxl human review export with masked account and fingerprint)

Wrapped into canonical ArtifactPayloads for atomic commitment via Nabhi.
"""

from __future__ import annotations

import io
from decimal import Decimal
from typing import Sequence

import openpyxl
import polars as pl
from openpyxl.styles import Font, PatternFill

from sarathi.sankalpa import ArtifactIntent, ArtifactPayload
from sarathi.shakti.bank_statements.models import (
    BankStatement,
    BankStatementConsolidationResult,
    ValidationStatus,
)


def consolidate_statements(statements: Sequence[BankStatement]) -> BankStatementConsolidationResult:
    """Consolidate multiple statements into a unified consolidation result in stable chronological order."""
    import datetime

    def _earliest_tx_date(stmt: BankStatement) -> datetime.date:
        dates = [tx.transaction_date for tx in stmt.transactions if tx.transaction_date is not None]
        return min(dates) if dates else datetime.date.min

    sorted_statements = sorted(statements, key=_earliest_tx_date)

    total_debits = Decimal("0")
    total_credits = Decimal("0")
    total_txns = 0
    overall_status = ValidationStatus.VALID
    all_issues = []

    for stmt in sorted_statements:
        if stmt.status == ValidationStatus.INVALID:
            overall_status = ValidationStatus.INVALID
        elif stmt.status == ValidationStatus.WARNING and overall_status == ValidationStatus.VALID:
            overall_status = ValidationStatus.WARNING
        all_issues.extend(stmt.issues)

        for tx in stmt.transactions:
            total_txns += 1
            if tx.debit is not None:
                total_debits += tx.debit
            if tx.credit is not None:
                total_credits += tx.credit

    return BankStatementConsolidationResult(
        statements=tuple(sorted_statements),
        total_transactions=total_txns,
        total_debit=total_debits,
        total_credit=total_credits,
        status=overall_status,
        issues=tuple(all_issues),
    )


def build_parquet_artifact(consolidation: BankStatementConsolidationResult) -> ArtifactPayload:
    """Generate Consolidated_Bank_Statement.parquet payload preserving exact Decimal precision."""
    dates: list[str] = []
    times: list[str | None] = []
    descriptions: list[str] = []
    ref_nums: list[str | None] = []
    chq_nums: list[str | None] = []
    debits: list[Decimal | None] = []
    credits: list[Decimal | None] = []
    balances: list[Decimal | None] = []
    bank_names: list[str] = []
    masked_accs: list[str | None] = []
    fingerprints: list[str | None] = []
    acc_holders: list[str | None] = []
    currencies: list[str] = []
    statuses: list[str] = []

    for stmt in consolidation.statements:
        ident = stmt.account_identity
        masked_acc = ident.masked_account_number if ident else None
        fingerprint = ident.account_fingerprint if ident else None
        holder = ident.account_holder if ident else None

        for tx in stmt.transactions:
            dates.append(tx.transaction_date.isoformat())
            times.append(tx.transaction_time.isoformat() if tx.transaction_time else None)
            descriptions.append(tx.description)
            ref_nums.append(tx.reference_number)
            chq_nums.append(tx.cheque_number)
            debits.append(tx.debit)
            credits.append(tx.credit)
            balances.append(tx.running_balance)
            bank_names.append(tx.bank_name)
            masked_accs.append(masked_acc)
            fingerprints.append(fingerprint)
            acc_holders.append(holder)
            currencies.append(tx.currency)
            statuses.append(tx.status.value)

    df = pl.DataFrame({
        "date": pl.Series("date", dates, dtype=pl.Utf8),
        "time": pl.Series("time", times, dtype=pl.Utf8),
        "description": pl.Series("description", descriptions, dtype=pl.Utf8),
        "reference_number": pl.Series("reference_number", ref_nums, dtype=pl.Utf8),
        "cheque_number": pl.Series("cheque_number", chq_nums, dtype=pl.Utf8),
        "debit": pl.Series("debit", debits, dtype=pl.Decimal(38, 2)),
        "credit": pl.Series("credit", credits, dtype=pl.Decimal(38, 2)),
        "running_balance": pl.Series("running_balance", balances, dtype=pl.Decimal(38, 2)),
        "bank_name": pl.Series("bank_name", bank_names, dtype=pl.Utf8),
        "masked_account_number": pl.Series("masked_account_number", masked_accs, dtype=pl.Utf8),
        "account_fingerprint": pl.Series("account_fingerprint", fingerprints, dtype=pl.Utf8),
        "account_holder": pl.Series("account_holder", acc_holders, dtype=pl.Utf8),
        "currency": pl.Series("currency", currencies, dtype=pl.Utf8),
        "status": pl.Series("status", statuses, dtype=pl.Utf8),
    })

    buf = io.BytesIO()
    df.write_parquet(buf)
    content_bytes = buf.getvalue()

    intent = ArtifactIntent(
        name="Consolidated_Bank_Statement.parquet",
        role="consolidated_data",
        media_type="application/vnd.apache.parquet",
    )
    return ArtifactPayload(intent=intent, content=content_bytes)


def build_xlsx_artifact(consolidation: BankStatementConsolidationResult) -> ArtifactPayload:
    """Generate Consolidated_Bank_Statement.xlsx payload with masked account and fingerprint."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Consolidated Statements"

    headers = [
        "Date", "Time", "Description", "Reference No.", "Cheque No.",
        "Debit", "Credit", "Running Balance", "Bank", "Masked Account", "Account Fingerprint", "Account Holder", "Status"
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font

    for stmt in consolidation.statements:
        ident = stmt.account_identity
        masked_acc = ident.masked_account_number if ident else ""
        fingerprint = ident.account_fingerprint if ident else ""
        holder = ident.account_holder if ident else ""

        for tx in stmt.transactions:
            ws.append([
                tx.transaction_date.strftime("%d-%m-%Y"),
                tx.transaction_time.strftime("%H:%M:%S") if tx.transaction_time else "",
                tx.description,
                tx.reference_number or "",
                tx.cheque_number or "",
                str(tx.debit) if tx.debit is not None else "",
                str(tx.credit) if tx.credit is not None else "",
                str(tx.running_balance) if tx.running_balance is not None else "",
                tx.bank_name,
                masked_acc or "",
                fingerprint or "",
                holder or "",
                tx.status.value.upper(),
            ])

    buf = io.BytesIO()
    wb.save(buf)
    content_bytes = buf.getvalue()

    intent = ArtifactIntent(
        name="Consolidated_Bank_Statement.xlsx",
        role="consolidated_report",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    return ArtifactPayload(intent=intent, content=content_bytes)
