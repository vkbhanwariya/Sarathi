"""Unit tests for Shruti — Read / Native Extraction capability."""

import struct
import sys
from pathlib import Path
from unittest.mock import patch

import openpyxl
import pymupdf
import pytest

from sarathi.dosh import DoshError, FailureCode
from sarathi.sankalpa import (
    CanonicalDocument,
    ExecutionContext,
    InputRef,
    Request,
    Result,
)
from sarathi.shakti.native_extraction import (
    CAPABILITY_DECLARATION,
    PLUGIN_INFO,
    NativeExtractionCapability,
)


@pytest.fixture
def capability() -> NativeExtractionCapability:
    return NativeExtractionCapability()


@pytest.fixture
def context() -> ExecutionContext:
    return ExecutionContext(
        run_id="run-1",
        request_id="req-1",
        trace_id="tr-1",
        span_id="sp-1",
    )


def _build_minimal_biff8_xls(sheet_name: str = "Accounts", cell_value: float = 999.5) -> bytes:
    """Build a genuine minimal valid BIFF8 .xls binary stream."""
    bof_wb = b"\x09\x08\x10\x00\x00\x06\x05\x00\xbb\x0d\xcc\x07\x00\x00\x00\x00\x00\x00\x00\x00"
    name_bytes = sheet_name.encode("ascii")
    data_len = 6 + 1 + 1 + len(name_bytes)
    sheet_bof_offset = 28 + data_len
    boundsheet = (
        struct.pack("<HHIBB", 0x0085, data_len, sheet_bof_offset, 0, 0)
        + struct.pack("B", len(name_bytes))
        + b"\x00"  # compressed ascii
        + name_bytes
    )
    eof_wb = b"\x0a\x00\x00\x00"

    bof_ws = b"\x09\x08\x10\x00\x00\x06\x10\x00\xbb\x0d\xcc\x07\x00\x00\x00\x00\x00\x00\x00\x00"
    num_rec0 = b"\x03\x02\x0e\x00" + struct.pack("<HHH", 0, 0, 0) + struct.pack("<d", 100.0)
    num_rec1 = b"\x03\x02\x0e\x00" + struct.pack("<HHH", 1, 0, 0) + struct.pack("<d", cell_value)
    eof_ws = b"\x0a\x00\x00\x00"

    return bof_wb + boundsheet + eof_wb + bof_ws + num_rec0 + num_rec1 + eof_ws


class TestNativeExtraction:
    def test_pdf_native_text_and_table_extraction(
        self, capability: NativeExtractionCapability, context: ExecutionContext, tmp_path: Path
    ) -> None:
        pdf_path = tmp_path / "sample.pdf"
        doc_pdf = pymupdf.open()
        page = doc_pdf.new_page(width=595, height=842)
        page.insert_text((72, 72), "Bank Account Statement\nAccount Number: 123456789")
        doc_pdf.save(str(pdf_path))
        doc_pdf.close()

        req = Request(
            request_id="req-1",
            requirement="read_native",
            inputs=(
                InputRef(
                    input_id="inp-pdf",
                    source_path=pdf_path,
                    display_name="sample.pdf",
                    size_bytes=pdf_path.stat().st_size,
                ),
            ),
        )

        res = capability.execute(req, context)
        assert isinstance(res, Result)
        assert res.next_requirement is None  # Usable text found -> no OCR needed

        doc = res.data
        assert isinstance(doc, CanonicalDocument)
        assert doc.detected_type == "pdf"
        assert len(doc.pages) == 1
        assert "Bank Account Statement" in doc.pages[0].text
        assert "123456789" in doc.pages[0].text
        assert len(doc.pages[0].spans) > 0

        # Provenance verification
        assert len(res.provenance) == 1
        prov = res.provenance[0]
        assert prov.source_input_id == "inp-pdf"
        assert prov.stage == "read_native"
        assert prov.capability_id == "read_native"
        assert prov.evidence["reader"] == "pymupdf"
        assert prov.source_file is None  # Zero path leakage in provenance

    def test_privacy_zero_raw_filesystem_path_leakage(
        self, capability: NativeExtractionCapability, context: ExecutionContext, tmp_path: Path
    ) -> None:
        # Create a document inside a secret directory path
        secret_dir = tmp_path / "very_secret_customer_pii_dir"
        secret_dir.mkdir(parents=True, exist_ok=True)
        doc_file = secret_dir / "statement.csv"
        doc_file.write_text("Header1,Header2\nValue1,Value2", encoding="utf-8")
        raw_path_str = str(doc_file)

        req = Request(
            request_id="req-privacy-1",
            requirement="read_native",
            inputs=(
                InputRef(
                    input_id="inp-priv",
                    source_path=doc_file,
                    display_name="statement.csv",
                    size_bytes=doc_file.stat().st_size,
                ),
            ),
        )

        res = capability.execute(req, context)

        # 1. Check provenance has no raw source path
        for prov in res.provenance:
            assert prov.source_file is None
            assert raw_path_str not in str(prov.evidence)
            assert str(secret_dir) not in str(prov.evidence)

        # 2. Check warnings have no raw source path
        for warn in res.warnings:
            assert raw_path_str not in warn.message
            assert str(secret_dir) not in warn.message

        # 3. Check document metadata has no raw source path
        doc = res.data
        assert raw_path_str not in str(doc.metadata)

        # 4. Check raised DoshError for missing file has no raw source path
        missing_req = Request(
            request_id="req-missing",
            requirement="read_native",
            inputs=(
                InputRef(
                    input_id="inp-missing",
                    source_path=secret_dir / "does_not_exist.pdf",
                    display_name="does_not_exist.pdf",
                    size_bytes=10,
                ),
            ),
        )
        with pytest.raises(DoshError) as exc_info:
            capability.execute(missing_req, context)

        err = exc_info.value
        assert raw_path_str not in err.message
        assert str(secret_dir) not in err.message

    def test_actual_bytes_override_misleading_extension(
        self, capability: NativeExtractionCapability, context: ExecutionContext, tmp_path: Path
    ) -> None:
        # Create an HTML table with an .xls extension
        misleading_file = tmp_path / "report.xls"
        html_content = """
        <html>
        <body>
            <table id="Transactions">
                <tr><th>Date</th><th>Description</th><th>Amount</th></tr>
                <tr><td>2026-01-01</td><td>Salary</td><td>50000</td></tr>
                <tr><td>2026-01-02</td><td>Rent</td><td>15000</td></tr>
            </table>
        </body>
        </html>
        """
        misleading_file.write_text(html_content, encoding="utf-8")

        req = Request(
            request_id="req-1",
            requirement="read_native",
            inputs=(
                InputRef(
                    input_id="inp-html-xls",
                    source_path=misleading_file,
                    display_name="report.xls",
                    size_bytes=misleading_file.stat().st_size,
                ),
            ),
        )

        res = capability.execute(req, context)
        assert res.next_requirement is None
        doc = res.data
        assert isinstance(doc, CanonicalDocument)
        assert doc.detected_type == "html_table"
        assert len(doc.tables) == 1
        table = doc.tables[0]
        assert table.name == "Transactions"
        assert table.headers == ("Date", "Description", "Amount")
        assert len(table.rows) == 2
        assert table.rows[0] == ("2026-01-01", "Salary", "50000")
        assert table.rows[1] == ("2026-01-02", "Rent", "15000")
        assert res.provenance[0].evidence["reader"] == "beautifulsoup4"

    def test_genuine_legacy_biff_xls_extraction(
        self, capability: NativeExtractionCapability, context: ExecutionContext, tmp_path: Path
    ) -> None:
        biff_file = tmp_path / "legacy.xls"
        biff_data = _build_minimal_biff8_xls("Accounts", 999.5)
        biff_file.write_bytes(biff_data)

        req = Request(
            request_id="req-biff",
            requirement="read_native",
            inputs=(
                InputRef(
                    input_id="inp-biff",
                    source_path=biff_file,
                    display_name="legacy.xls",
                    size_bytes=biff_file.stat().st_size,
                ),
            ),
        )

        res = capability.execute(req, context)
        assert res.next_requirement is None
        doc = res.data
        assert isinstance(doc, CanonicalDocument)
        assert doc.detected_type == "xls_legacy"
        assert len(doc.tables) == 1
        assert doc.tables[0].name == "Accounts"
        assert doc.tables[0].rows == ((999.5,),)
        assert res.provenance[0].evidence["reader"] == "xlrd"

    def test_xlsm_macro_workbook_extraction(
        self, capability: NativeExtractionCapability, context: ExecutionContext, tmp_path: Path
    ) -> None:
        xlsm_path = tmp_path / "macro_enabled.xlsm"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "MacroData"
        ws.append(["Param", "Value"])
        ws.append(["Threshold", 500])
        wb.save(str(xlsm_path))
        wb.close()

        req = Request(
            request_id="req-xlsm",
            requirement="read_native",
            inputs=(
                InputRef(
                    input_id="inp-xlsm",
                    source_path=xlsm_path,
                    display_name="macro_enabled.xlsm",
                    size_bytes=xlsm_path.stat().st_size,
                ),
            ),
        )

        res = capability.execute(req, context)
        assert res.next_requirement is None
        doc = res.data
        assert isinstance(doc, CanonicalDocument)
        assert doc.detected_type == "xlsx"
        assert len(doc.tables) == 1
        assert doc.tables[0].name == "MacroData"
        assert doc.tables[0].headers == ("Param", "Value")
        assert doc.tables[0].rows == (("Threshold", 500),)

    def test_xlsx_multi_sheet_extraction_retains_all_sheets(
        self, capability: NativeExtractionCapability, context: ExecutionContext, tmp_path: Path
    ) -> None:
        xlsx_path = tmp_path / "multi_sheet.xlsx"
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Summary"
        ws1.append(["Metric", "Value"])
        ws1.append(["Total", 1000])

        ws2 = wb.create_sheet(title="Details")
        ws2.append(["ID", "Item", "Cost"])
        ws2.append([1, "Hardware", 400])
        ws2.append([2, "Software", 600])

        wb.save(str(xlsx_path))
        wb.close()

        req = Request(
            request_id="req-1",
            requirement="read_native",
            inputs=(
                InputRef(
                    input_id="inp-xlsx",
                    source_path=xlsx_path,
                    display_name="multi_sheet.xlsx",
                    size_bytes=xlsx_path.stat().st_size,
                ),
            ),
        )

        res = capability.execute(req, context)
        assert res.next_requirement is None
        doc = res.data
        assert isinstance(doc, CanonicalDocument)
        assert doc.detected_type == "xlsx"

        # Verify ALL sheets are present, not just the first or largest
        assert len(doc.tables) == 2
        table_names = [t.name for t in doc.tables]
        assert "Summary" in table_names
        assert "Details" in table_names

        summary_tab = next(t for t in doc.tables if t.name == "Summary")
        assert summary_tab.headers == ("Metric", "Value")
        assert summary_tab.rows[0] == ("Total", 1000)

        details_tab = next(t for t in doc.tables if t.name == "Details")
        assert details_tab.headers == ("ID", "Item", "Cost")
        assert len(details_tab.rows) == 2

        # Provenance for each sheet
        assert len(res.provenance) == 2
        sheets_in_prov = [p.evidence["sheet"] for p in res.provenance]
        assert "Summary" in sheets_in_prov
        assert "Details" in sheets_in_prov

    def test_csv_encoding_recovery_and_parsing(
        self, capability: NativeExtractionCapability, context: ExecutionContext, tmp_path: Path
    ) -> None:
        csv_path = tmp_path / "latin1.csv"
        # Write CSV encoded in latin-1 with special characters
        content = "Namn;Stad;Belopp\nMüller;München;1200\nAndré;Genève;3400\n"
        csv_path.write_bytes(content.encode("iso-8859-1"))

        req = Request(
            request_id="req-1",
            requirement="read_native",
            inputs=(
                InputRef(
                    input_id="inp-csv",
                    source_path=csv_path,
                    display_name="latin1.csv",
                    size_bytes=csv_path.stat().st_size,
                ),
            ),
        )

        res = capability.execute(req, context)
        assert res.next_requirement is None
        doc = res.data
        assert isinstance(doc, CanonicalDocument)
        assert doc.detected_type == "csv_or_text"
        assert len(doc.tables) == 1
        table = doc.tables[0]
        assert "Namn" in table.headers or "Müller" in str(table.rows)
        assert "München" in str(table.rows)

    def test_spreadsheet_ml_xml_extraction(
        self, capability: NativeExtractionCapability, context: ExecutionContext, tmp_path: Path
    ) -> None:
        xml_path = tmp_path / "spreadsheet.xml"
        xml_content = """<?xml version="1.0"?>
        <?mso-application progid="Excel.Sheet"?>
        <Workbook xmlns="urn:schemas-microsoft-com:office:spreadsheet"
         xmlns:ss="urn:schemas-microsoft-com:office:spreadsheet">
         <Worksheet ss:Name="TaxReport">
          <Table>
           <Row>
            <Cell><Data ss:Type="String">Category</Data></Cell>
            <Cell><Data ss:Type="String">Amount</Data></Cell>
           </Row>
           <Row>
            <Cell><Data ss:Type="String">GST</Data></Cell>
            <Cell><Data ss:Type="Number">1800</Data></Cell>
           </Row>
          </Table>
         </Worksheet>
        </Workbook>"""
        xml_path.write_text(xml_content, encoding="utf-8")

        req = Request(
            request_id="req-1",
            requirement="read_native",
            inputs=(
                InputRef(
                    input_id="inp-xml",
                    source_path=xml_path,
                    display_name="spreadsheet.xml",
                    size_bytes=xml_path.stat().st_size,
                ),
            ),
        )

        res = capability.execute(req, context)
        assert res.next_requirement is None
        doc = res.data
        assert isinstance(doc, CanonicalDocument)
        assert doc.detected_type == "spreadsheet_ml"
        assert len(doc.tables) == 1
        table = doc.tables[0]
        assert table.name == "TaxReport"
        assert table.headers == ("Category", "Amount")
        assert table.rows == (("GST", "1800"),)

    def test_empty_native_output_requests_ocr(
        self, capability: NativeExtractionCapability, context: ExecutionContext, tmp_path: Path
    ) -> None:
        # Create a PDF with a blank page (no text stream)
        blank_pdf_path = tmp_path / "scanned_blank.pdf"
        doc_pdf = pymupdf.open()
        doc_pdf.new_page(width=595, height=842)
        doc_pdf.save(str(blank_pdf_path))
        doc_pdf.close()

        req = Request(
            request_id="req-1",
            requirement="read_native",
            inputs=(
                InputRef(
                    input_id="inp-blank",
                    source_path=blank_pdf_path,
                    display_name="scanned_blank.pdf",
                    size_bytes=blank_pdf_path.stat().st_size,
                ),
            ),
        )

        res = capability.execute(req, context)
        # Empty text must escalate to OCR
        assert res.next_requirement == "ocr"
        assert any(w.code == "NATIVE_EXTRACTION_EMPTY" for w in res.warnings)

    def test_corrupted_file_requests_ocr(
        self, capability: NativeExtractionCapability, context: ExecutionContext, tmp_path: Path
    ) -> None:
        corrupt_path = tmp_path / "corrupt.pdf"
        corrupt_path.write_bytes(b"%PDF-1.4\ncorrupt damaged stream %%%")

        req = Request(
            request_id="req-1",
            requirement="read_native",
            inputs=(
                InputRef(
                    input_id="inp-corrupt",
                    source_path=corrupt_path,
                    display_name="corrupt.pdf",
                    size_bytes=corrupt_path.stat().st_size,
                ),
            ),
        )

        res = capability.execute(req, context)
        # Corrupted input must escalate to OCR
        assert res.next_requirement == "ocr"

    def test_multi_input_one_native_one_ocr_required(
        self, capability: NativeExtractionCapability, context: ExecutionContext, tmp_path: Path
    ) -> None:
        # Input 1: valid native PDF
        valid_pdf_path = tmp_path / "valid.pdf"
        doc_pdf = pymupdf.open()
        p = doc_pdf.new_page()
        p.insert_text((50, 50), "Valid invoice text 5000 USD")
        doc_pdf.save(str(valid_pdf_path))
        doc_pdf.close()

        # Input 2: blank scanned PDF (no native text)
        blank_pdf_path = tmp_path / "blank.pdf"
        doc_blank = pymupdf.open()
        doc_blank.new_page()
        doc_blank.save(str(blank_pdf_path))
        doc_blank.close()

        req = Request(
            request_id="req-multi",
            requirement="read_native",
            inputs=(
                InputRef(
                    input_id="inp-valid",
                    source_path=valid_pdf_path,
                    display_name="valid.pdf",
                    size_bytes=valid_pdf_path.stat().st_size,
                ),
                InputRef(
                    input_id="inp-blank",
                    source_path=blank_pdf_path,
                    display_name="blank.pdf",
                    size_bytes=blank_pdf_path.stat().st_size,
                ),
            ),
        )

        res = capability.execute(req, context)

        # Since input 2 requires OCR, next_requirement is signaled
        assert res.next_requirement == "ocr"

        # Result data preserves BOTH document outputs
        assert isinstance(res.data, tuple)
        assert len(res.data) == 2

        doc1, doc2 = res.data
        assert isinstance(doc1, CanonicalDocument)
        assert doc1.source_input_id == "inp-valid"
        assert "Valid invoice text 5000 USD" in doc1.pages[0].text

        assert isinstance(doc2, CanonicalDocument)
        assert doc2.source_input_id == "inp-blank"

        # Both source input IDs have their respective provenance
        prov_input_ids = [p.source_input_id for p in res.provenance]
        assert "inp-valid" in prov_input_ids
        assert "inp-blank" in prov_input_ids

    def test_unexpected_code_errors_are_not_caught_as_ocr_handoff(
        self, capability: NativeExtractionCapability, context: ExecutionContext, tmp_path: Path
    ) -> None:
        # Create a valid PDF
        pdf_path = tmp_path / "test.pdf"
        doc_pdf = pymupdf.open()
        doc_pdf.new_page().insert_text((50, 50), "Test")
        doc_pdf.save(str(pdf_path))
        doc_pdf.close()

        req = Request(
            request_id="req-bug",
            requirement="read_native",
            inputs=(
                InputRef(
                    input_id="inp-1",
                    source_path=pdf_path,
                    display_name="test.pdf",
                    size_bytes=pdf_path.stat().st_size,
                ),
            ),
        )

        # Simulate an unexpected programmer/code bug (e.g. AttributeError) inside reader
        with patch("sarathi.shakti.native_extraction.capability.read_pdf", side_effect=AttributeError("Bug in code")):
            with pytest.raises(AttributeError, match="Bug in code"):
                capability.execute(req, context)

    def test_forced_value_error_does_not_become_ocr_requirement(
        self, capability: NativeExtractionCapability, context: ExecutionContext, tmp_path: Path
    ) -> None:
        pdf_path = tmp_path / "test_val.pdf"
        doc_pdf = pymupdf.open()
        doc_pdf.new_page().insert_text((50, 50), "Sample Text")
        doc_pdf.save(str(pdf_path))
        doc_pdf.close()

        req = Request(
            request_id="req-val-err",
            requirement="read_native",
            inputs=(
                InputRef(
                    input_id="inp-val",
                    source_path=pdf_path,
                    display_name="test_val.pdf",
                    size_bytes=pdf_path.stat().st_size,
                ),
            ),
        )

        with patch(
            "sarathi.shakti.native_extraction.capability.read_pdf", side_effect=ValueError("Contract violation")
        ):
            with pytest.raises(ValueError, match="Contract violation"):
                capability.execute(req, context)

    def test_unexpected_xlsx_reader_failure_does_not_fall_back_or_request_ocr(
        self, capability: NativeExtractionCapability, context: ExecutionContext, tmp_path: Path
    ) -> None:
        xlsx_path = tmp_path / "test_xlsx.xlsx"
        wb = openpyxl.Workbook()
        wb.active.append(["Header", "Value"])
        wb.save(str(xlsx_path))
        wb.close()

        req = Request(
            request_id="req-xlsx-err",
            requirement="read_native",
            inputs=(
                InputRef(
                    input_id="inp-xlsx",
                    source_path=xlsx_path,
                    display_name="test_xlsx.xlsx",
                    size_bytes=xlsx_path.stat().st_size,
                ),
            ),
        )

        # Programmer/system bug in Calamine reader must not fall back to openpyxl or request OCR
        with patch("python_calamine.CalamineWorkbook.from_object", side_effect=TypeError("Unexpected type bug")):
            with pytest.raises(TypeError, match="Unexpected type bug"):
                capability.execute(req, context)

    def test_pdf_table_detection_known_error_emits_warning_and_preserves_text(
        self, capability: NativeExtractionCapability, context: ExecutionContext, tmp_path: Path
    ) -> None:
        pdf_path = tmp_path / "table_err.pdf"
        doc_pdf = pymupdf.open()
        p = doc_pdf.new_page()
        p.insert_text((50, 50), "Important statement body text")
        doc_pdf.save(str(pdf_path))
        doc_pdf.close()

        req = Request(
            request_id="req-tab-err",
            requirement="read_native",
            inputs=(
                InputRef(
                    input_id="inp-tab",
                    source_path=pdf_path,
                    display_name="table_err.pdf",
                    size_bytes=pdf_path.stat().st_size,
                ),
            ),
        )

        # Simulate a known table detection issue on find_tables
        with patch.object(pymupdf.Page, "find_tables", side_effect=ValueError("Malformed vector graphics")):
            res = capability.execute(req, context)

        assert res.next_requirement is None
        doc = res.data
        assert isinstance(doc, CanonicalDocument)
        assert "Important statement body text" in doc.pages[0].text
        assert any(w.code == "PDF_TABLE_DETECTION_SKIPPED" for w in res.warnings)

    def test_unexpected_error_during_pdf_table_conversion_propagates(
        self, capability: NativeExtractionCapability, context: ExecutionContext, tmp_path: Path
    ) -> None:
        from unittest.mock import MagicMock

        pdf_path = tmp_path / "table_conv_err.pdf"
        doc_pdf = pymupdf.open()
        p = doc_pdf.new_page()
        p.insert_text((50, 50), "PDF with table")
        doc_pdf.save(str(pdf_path))
        doc_pdf.close()

        req = Request(
            request_id="req-tab-conv",
            requirement="read_native",
            inputs=(
                InputRef(
                    input_id="inp-tab-conv",
                    source_path=pdf_path,
                    display_name="table_conv_err.pdf",
                    size_bytes=pdf_path.stat().st_size,
                ),
            ),
        )

        mock_tab = MagicMock()
        mock_tab.extract.side_effect = ValueError("Unexpected table conversion bug")
        mock_finder = MagicMock()
        mock_finder.tables = [mock_tab]

        # Force an unexpected ValueError during table row extraction/conversion
        with patch.object(pymupdf.Page, "find_tables", return_value=mock_finder):
            with pytest.raises(ValueError, match="Unexpected table conversion bug"):
                capability.execute(req, context)

    def test_unexpected_zip_detector_error_propagates(self) -> None:
        from sarathi.shakti.native_extraction.detector import detect_content_format

        zip_data = b"PK\x03\x04\x14\x00\x00\x00\x08\x00some_zip_data"
        with patch("zipfile.ZipFile", side_effect=TypeError("Unexpected zipfile error")):
            with pytest.raises(TypeError, match="Unexpected zipfile error"):
                detect_content_format(zip_data)

    def test_oserror_during_source_read_becomes_safe_dosh_error(
        self, capability: NativeExtractionCapability, context: ExecutionContext, tmp_path: Path
    ) -> None:
        file_path = tmp_path / "read_err.pdf"
        file_path.write_bytes(b"%PDF-1.4\n")

        req = Request(
            request_id="req-io-err",
            requirement="read_native",
            inputs=(
                InputRef(
                    input_id="inp-io",
                    source_path=file_path,
                    display_name="read_err.pdf",
                    size_bytes=file_path.stat().st_size,
                ),
            ),
        )

        with patch.object(Path, "read_bytes", side_effect=PermissionError("Access denied to file")):
            with pytest.raises(DoshError) as exc_info:
                capability.execute(req, context)

        err = exc_info.value
        assert err.code is FailureCode.EXECUTION_FAILED
        assert "Failed to read source input file." in err.message
        assert "Access denied" not in err.message
        assert str(file_path) not in err.message

    def test_unknown_binary_content_returns_controlled_unsupported_error(
        self, capability: NativeExtractionCapability, context: ExecutionContext, tmp_path: Path
    ) -> None:
        bin_path = tmp_path / "binary.bin"
        bin_path.write_bytes(b"\x7fELF\x02\x01\x01\x00\x00\x00\x00\x00\x00\x00\x00\x00\x02\x00>\x00")

        req = Request(
            request_id="req-1",
            requirement="read_native",
            inputs=(
                InputRef(
                    input_id="inp-bin",
                    source_path=bin_path,
                    display_name="binary.bin",
                    size_bytes=bin_path.stat().st_size,
                ),
            ),
        )

        with pytest.raises(DoshError) as exc_info:
            capability.execute(req, context)

        assert exc_info.value.code is FailureCode.UNSUPPORTED

    def test_no_ocr_import_and_no_fabricated_confidence(
        self, capability: NativeExtractionCapability, context: ExecutionContext, tmp_path: Path
    ) -> None:
        txt_path = tmp_path / "test.txt"
        txt_path.write_text("Plain text sample", encoding="utf-8")

        req = Request(
            request_id="req-1",
            requirement="read_native",
            inputs=(
                InputRef(
                    input_id="inp-txt",
                    source_path=txt_path,
                    display_name="test.txt",
                    size_bytes=txt_path.stat().st_size,
                ),
            ),
        )

        res = capability.execute(req, context)
        # Confidence must be unavailable (None), not fabricated
        assert res.confidence is None

        # Verify no OCR engine modules have been imported
        forbidden_modules = {"rapidocr", "pytesseract", "tesseract", "easyocr", "paddleocr"}
        for mod in forbidden_modules:
            assert mod not in sys.modules

    def test_plugin_and_capability_declarations(self) -> None:
        assert PLUGIN_INFO.plugin_id == "shakti.native_extraction"
        assert "read_native" in PLUGIN_INFO.capabilities
        assert CAPABILITY_DECLARATION.capability_id == "read_native"
        assert CAPABILITY_DECLARATION.plugin_id == "shakti.native_extraction"
