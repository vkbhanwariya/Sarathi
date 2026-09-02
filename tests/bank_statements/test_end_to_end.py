"""End-to-End Operational Acceptance Test for Bank Statement Consolidation."""

import io
import json
from datetime import date
from decimal import Decimal
from pathlib import Path

import openpyxl
import polars as pl
import pytest

from sarathi.agni import Agni
from sarathi.darpana import Darpana
from sarathi.dosh import DoshError, FailureCode
from sarathi.sankalpa import (
    ExecutionContext,
    ExecutionProfile,
    InputRef,
    Request,
    Result,
)
from sarathi.shakti.bank_statements.capability import _CANONICAL_BANKS_DIR, BankStatementCapability
from sarathi.shakti.bank_statements.consolidator import (
    build_parquet_artifact,
    consolidate_statements,
)
from sarathi.shakti.bank_statements.models import (
    BankStatement,
    BankStatementConsolidationResult,
    Transaction,
    create_account_identity,
)

_SBI_FIXTURE = Path(__file__).parent / "fixtures" / "sbi_statement.csv"


def test_bank_capability_does_not_instantiate_native_extraction() -> None:
    """Verify BankStatementCapability does not own or instantiate NativeExtractionCapability."""
    cap = BankStatementCapability()
    assert not hasattr(cap, "_native_extractor")


def test_missing_prior_canonical_document_fails_safely() -> None:
    """Verify executing BankStatementCapability without prior CanonicalDocument fails safely with DoshError."""
    cap = BankStatementCapability()
    req = Request(
        request_id="req-fail-1",
        requirement="bank_statements",
        inputs=(InputRef("i1", _SBI_FIXTURE, "sbi.csv", 100),),
    )
    ctx = ExecutionContext("run-1", "req-fail-1", "t1", "s1")

    with pytest.raises(DoshError) as exc_info:
        cap.execute(req, ctx, prior_result=None)
    assert exc_info.value.code == FailureCode.VALIDATION_FAILED


def test_account_identity_deterministic_fingerprint_and_masking() -> None:
    """Verify AccountIdentity creates deterministic fingerprints and masks leading account digits."""
    ident1 = create_account_identity(
        bank_name="State Bank of India",
        raw_account_number="30123456789",
        account_holder="Rahul Sharma",
    )
    ident2 = create_account_identity(
        bank_name="State Bank of India",
        raw_account_number="30123456789",
        account_holder="Rahul Sharma",
    )
    ident_diff = create_account_identity(
        bank_name="State Bank of India",
        raw_account_number="98765432100",
        account_holder="Rahul Sharma",
    )

    assert ident1.masked_account_number == "XXXXXXX6789"
    assert ident1.account_fingerprint is not None
    assert len(ident1.account_fingerprint) == 16
    assert ident1.account_fingerprint == ident2.account_fingerprint
    assert ident1.account_fingerprint != ident_diff.account_fingerprint


def test_financial_decimal_values_roundtrip_exactly_in_parquet() -> None:
    """Verify financial decimal values like 0.10 and 1234567890.01 roundtrip exactly in Parquet without float errors."""
    ident = create_account_identity("Test Bank", "12345678")
    t1 = Transaction(
        transaction_date=date(2026, 1, 1),
        description="Micro debit",
        bank_name="Test Bank",
        debit=Decimal("0.10"),
        running_balance=Decimal("1234567890.01"),
        account_identity=ident,
    )
    stmt = BankStatement(
        bank_name="Test Bank",
        bank_profile="generic",
        account_identity=ident,
        transactions=(t1,),
    )
    consolidation = consolidate_statements([stmt])
    payload = build_parquet_artifact(consolidation)

    df = pl.read_parquet(io.BytesIO(payload.content))
    assert df["debit"].to_list() == [Decimal("0.10")]
    assert df["running_balance"].to_list() == [Decimal("1234567890.01")]


def test_canonical_banks_dir_resolution() -> None:
    """Verify bank configuration directory resolves canonically without machine-specific hardcoded paths."""
    assert _CANONICAL_BANKS_DIR.exists()
    assert (_CANONICAL_BANKS_DIR / "common.yaml").exists()
    assert (_CANONICAL_BANKS_DIR / "sbi.yaml").exists()
    assert (_CANONICAL_BANKS_DIR / "hdfc.yaml").exists()


def test_e2e_sbi_bank_statement_consolidation(tmp_path: Path) -> None:
    """Test full E2E execution through Agni composition root for Bank Statement Consolidation."""
    runtime_dir = tmp_path / "Runtime"
    output_dir = tmp_path / "Output"
    darpana = Darpana(capacity=500)

    agni = Agni(
        runtime_root=runtime_dir,
        output_root=output_dir,
        darpana=darpana,
    )

    inp = InputRef(
        input_id="inp-sbi-1",
        source_path=_SBI_FIXTURE,
        display_name="sbi_statement.csv",
        size_bytes=_SBI_FIXTURE.stat().st_size,
    )

    req = Request(
        request_id="req-sbi-e2e",
        requirement="bank_statements",
        inputs=(inp,),
        profile=ExecutionProfile.ACCURATE,
    )

    ctx = ExecutionContext(
        run_id="run-sbi-e2e", request_id="req-sbi-e2e", trace_id="trace-sbi-e2e", span_id="span-sbi-e2e"
    )

    # Execute through canonical Agni path (read_native -> bank_statements)
    result = agni.execute(req, context=ctx)

    # 1. Result verification
    assert isinstance(result, Result)
    assert isinstance(result.data, BankStatementConsolidationResult)
    consolidation: BankStatementConsolidationResult = result.data

    assert len(consolidation.statements) == 1
    stmt = consolidation.statements[0]
    assert stmt.bank_profile == "sbi"
    assert stmt.bank_name == "State Bank of India"
    assert stmt.account_identity is not None
    assert stmt.account_identity.masked_account_number == "XXXXXXX6789"
    assert stmt.account_identity.account_fingerprint is not None
    assert len(stmt.transactions) == 3  # 3 genuine transactions (UPI, Salary, ATM)
    assert consolidation.total_debit == Decimal("2050.00")
    assert consolidation.total_credit == Decimal("50000.00")

    # 2. Confirmed Artifacts verification in Result
    assert len(result.artifacts) == 2
    parquet_art = next((a for a in result.artifacts if a.role == "consolidated_data"), None)
    xlsx_art = next((a for a in result.artifacts if a.role == "consolidated_report"), None)

    assert parquet_art is not None
    assert parquet_art.path.exists()
    assert parquet_art.path.name == "Consolidated_Bank_Statement.parquet"
    assert parquet_art.size_bytes > 0

    assert xlsx_art is not None
    assert xlsx_art.path.exists()
    assert xlsx_art.path.name == "Consolidated_Bank_Statement.xlsx"
    assert xlsx_art.size_bytes > 0

    # 3. Verify PII Protection in Parquet & XLSX (Raw account number is absent)
    parquet_df = pl.read_parquet(parquet_art.path)
    assert "30123456789" not in str(parquet_df.to_dicts())
    assert parquet_df["masked_account_number"].to_list() == ["XXXXXXX6789", "XXXXXXX6789", "XXXXXXX6789"]
    assert parquet_df["debit"].to_list() == [Decimal("50.00"), None, Decimal("2000.00")]
    assert parquet_df["credit"].to_list() == [None, Decimal("50000.00"), None]

    wb = openpyxl.load_workbook(xlsx_art.path)
    ws = wb.active
    all_cell_values = [str(cell.value) for row in ws.iter_rows() for cell in row if cell.value is not None]
    assert "30123456789" not in all_cell_values
    assert "XXXXXXX6789" in all_cell_values

    # 4. Final Manifest verification
    manifest_path = parquet_art.path.parent / "run-manifest.json"
    assert manifest_path.exists()
    manifest_data = json.loads(manifest_path.read_text(encoding="utf-8"))
    assert manifest_data["status"] == "completed"
    assert manifest_data["run_id"] == ctx.run_id
    assert len(manifest_data["artifacts"]) == 2

    # 5. Telemetry verification in shared Darpana
    maruti_recs = tuple(r for r in darpana.maruti_records() if r.run_id == ctx.run_id)
    assert len(maruti_recs) > 0
    assert any(r.phase_name == "capability_execution" and r.outcome == "success" for r in maruti_recs)


def test_excel_bank_statement_with_metadata_headers(tmp_path: Path) -> None:
    runtime_dir = tmp_path / "Runtime"
    output_dir = tmp_path / "Output"
    darpana = Darpana(capacity=200)
    agni = Agni(runtime_root=runtime_dir, output_root=output_dir, darpana=darpana)

    xlsx_path = tmp_path / "sbi_statement.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Statement"
    # Metadata rows at the top (typical in real bank statements)
    ws.append(["STATE BANK OF INDIA"])
    ws.append(["Account Statement for Account Number: 30123456789"])
    ws.append(["Account Holder: Ramesh Kumar", "CIF No: 88776655"])
    ws.append(["Branch: Bhopal Main Branch", "IFSC: SBIN0001234"])
    ws.append([])  # blank row
    # Transaction table headers at row 6
    ws.append(["Txn Date", "Description", "Ref No", "Debit", "Credit", "Balance"])
    ws.append(["01/01/2026", "OPENING BALANCE", "", "", "", "10000.00"])
    ws.append(["05/01/2026", "ATM CASH WITHDRAWAL", "W123", "2000.00", "", "8000.00"])
    ws.append(["10/01/2026", "SALARY CREDIT", "C456", "", "50000.00", "58000.00"])
    wb.save(str(xlsx_path))
    wb.close()

    inp = InputRef(
        input_id="inp-sbi-xlsx",
        source_path=xlsx_path,
        display_name="sbi_statement.xlsx",
        size_bytes=xlsx_path.stat().st_size,
    )
    req = Request(
        request_id="req-sbi-xlsx",
        requirement="bank_statements",
        inputs=(inp,),
        profile=ExecutionProfile.INSTANT,
    )
    ctx = ExecutionContext("run-sbi-xlsx", "req-sbi-xlsx", "t-sbi", "s-sbi")
    result = agni.execute(req, context=ctx)
    assert isinstance(result, Result)
    consolidation: BankStatementConsolidationResult = result.data
    assert consolidation.total_transactions == 2
